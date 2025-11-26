from werkzeug.utils import secure_filename
import uuid
from conexion.conexionBD import connectionBD
import datetime
import re
import os
from os import remove
from os import path
import openpyxl
from flask import send_file


def procesar_form_empleado(dataForm, foto_perfil):
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    salario_entero = int(salario_sin_puntos)
    result_foto_perfil = procesar_imagen_perfil(foto_perfil)

    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """
                    INSERT INTO tbl_empleados (
                        nombre_empleado,
                        apellido_empleado,
                        sexo_empleado,
                        telefono_empleado,
                        email_empleado,
                        profesion_empleado,
                        foto_empleado,
                        salario_empleado
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """

                valores = (
                    dataForm['nombre_empleado'],
                    dataForm['apellido_empleado'],
                    dataForm['sexo_empleado'],
                    dataForm['telefono_empleado'],
                    dataForm['email_empleado'],
                    dataForm['profesion_empleado'],
                    result_foto_perfil,
                    salario_entero
                )

                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return cursor.rowcount

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, '../static/fotos_empleados/')

        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            os.chmod(upload_dir, 0o755)

        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_empleado,
                        nombre_empleado,
                        apellido_empleado,
                        salario_empleado,
                        foto_empleado,
                        CASE
                            WHEN sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados
                    ORDER BY id_empleado DESC
                """
                cursor.execute(querySQL)
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en la función sql_lista_empleadosBD: {e}")
        return None


def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_empleado,
                        nombre_empleado,
                        apellido_empleado,
                        salario_empleado,
                        CASE
                            WHEN sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        telefono_empleado,
                        email_empleado,
                        profesion_empleado,
                        foto_empleado,
                        DATE_FORMAT(fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados
                    WHERE id_empleado = %s
                """
                cursor.execute(querySQL, (idEmpleado,))
                return cursor.fetchone()
    except Exception as e:
        print(f"Error en la función sql_detalles_empleadosBD: {e}")
        return None


def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_empleado,
                        nombre_empleado,
                        apellido_empleado,
                        salario_empleado,
                        email_empleado,
                        telefono_empleado,
                        profesion_empleado,
                        DATE_FORMAT(fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados
                    ORDER BY id_empleado DESC
                """
                cursor.execute(querySQL)
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    cabeceraExcel = (
        "Nombre", "Apellido", "Sexo",
        "Telefono", "Email", "Profesión",
        "Salario", "Fecha de Ingreso"
    )
    hoja.append(cabeceraExcel)

    formato_moneda = '#,##0'

    for r in dataEmpleados:
        hoja.append((
            r['nombre_empleado'],
            r['apellido_empleado'],
            r['sexo_empleado'],
            r['telefono_empleado'],
            r['email_empleado'],
            r['profesion_empleado'],
            r['salario_empleado'],
            r['fecha_registro']
        ))

    for fila_num in range(2, hoja.max_row + 1):
        celda = hoja.cell(row=fila_num, column=7)
        celda.number_format = formato_moneda

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)

def inventarioReporte():
    """Obtiene los datos del inventario para el reporte en Excel."""
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        nombre,
                        unidad,
                        descripcion,
                        categoria,
                        stock,
                        stock_minimo,
                        estado_stock,
                        precio_compra,
                        precio_venta,
                        DATE_FORMAT(fecha_registro, '%d-%m-%Y %H:%i') AS fecha_registro
                    FROM productos
                    ORDER BY id DESC
                """
                cursor.execute(querySQL)
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en la función inventarioReporte: {e}")
        return []


def generarReporteExcelInventario():
    """Genera y devuelve el archivo Excel del inventario."""
    dataProductos = inventarioReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    cabeceraExcel = (
        "Nombre",
        "Unidad",
        "Descripción",
        "Categoría",
        "Stock",
        "Stock mínimo",
        "Estado stock",
        "Precio compra",
        "Precio venta",
        "Fecha de registro",
    )
    hoja.append(cabeceraExcel)

    # Formato de moneda para precios
    formato_moneda = '#,##0.00'

    for p in dataProductos:
        hoja.append((
            p['nombre'],
            p['unidad'],
            p['descripcion'],
            p['categoria'],
            p['stock'],
            p['stock_minimo'],
            p['estado_stock'],
            p['precio_compra'],
            p['precio_venta'],
            p['fecha_registro'],
        ))

    # Aplicar formato moneda a columnas de precios (8 y 9)
    for fila_num in range(2, hoja.max_row + 1):
        celda_compra = hoja.cell(row=fila_num, column=8)
        celda_venta = hoja.cell(row=fila_num, column=9)
        celda_compra.number_format = formato_moneda
        celda_venta.number_format = formato_moneda

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_inventario_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)



def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_empleado,
                        nombre_empleado,
                        apellido_empleado,
                        salario_empleado,
                        CASE
                            WHEN sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados
                    WHERE nombre_empleado LIKE %s
                    ORDER BY id_empleado DESC
                """
                pattern = f"%{search}%"
                cursor.execute(querySQL, (pattern,))
                return cursor.fetchall()
    except Exception as e:
        print(f"Ocurrió un error en buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_empleado,
                        nombre_empleado,
                        apellido_empleado,
                        sexo_empleado,
                        telefono_empleado,
                        email_empleado,
                        profesion_empleado,
                        salario_empleado,
                        foto_empleado
                    FROM tbl_empleados
                    WHERE id_empleado = %s
                    LIMIT 1
                """
                cursor.execute(querySQL, (id,))
                return cursor.fetchone()
    except Exception as e:
        print(f"Ocurrió un error en buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                salario_sin_puntos = re.sub('[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)

                query = """
                    UPDATE tbl_empleados
                    SET 
                        nombre_empleado = %s,
                        apellido_empleado = %s,
                        sexo_empleado = %s,
                        telefono_empleado = %s,
                        email_empleado = %s,
                        profesion_empleado = %s,
                        salario_empleado = %s
                """

                params = [
                    data.form['nombre_empleado'],
                    data.form['apellido_empleado'],
                    data.form['sexo_empleado'],
                    data.form['telefono_empleado'],
                    data.form['email_empleado'],
                    data.form['profesion_empleado'],
                    salario_empleado
                ]

                if 'foto_empleado' in data.files and data.files['foto_empleado'].filename != '':
                    fotoNueva = procesar_imagen_perfil(data.files['foto_empleado'])
                    query += ", foto_empleado = %s"
                    params.append(fotoNueva)

                query += " WHERE id_empleado = %s"
                params.append(data.form['id_empleado'])

                cursor.execute(query, params)
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# =========================
#   INVENTARIO
# =========================

def sql_lista_inventarioBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id,
                        nombre,
                        unidad,
                        descripcion,
                        categoria,
                        stock,
                        stock_minimo,
                        estado_stock,
                        precio_compra,
                        precio_venta,
                        fecha_registro
                    FROM productos
                    ORDER BY id DESC
                """
                cursor.execute(querySQL)
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en sql_lista_inventarioBD: {e}")
        return []


def sql_detalle_productoBD(id_producto):
    """Detalle de un producto para la vista /detalles-producto/<id>."""
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id,
                        nombre,
                        unidad,
                        descripcion,
                        categoria,
                        stock,
                        stock_minimo,
                        estado_stock,
                        precio_compra,
                        precio_venta,
                        fecha_registro
                    FROM productos
                    WHERE id = %s
                    LIMIT 1
                """
                cursor.execute(querySQL, (id_producto,))
                return cursor.fetchone()
    except Exception as e:
        print(f"Error en sql_detalle_productoBD: {e}")
        return None


def buscarProductoInventarioUnico(id_producto):
    """Trae un solo producto para el formulario de actualización."""
    return sql_detalle_productoBD(id_producto)


def procesar_actualizacion_producto(data):
    """Procesa el form de actualización de un producto."""
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                query = """
                    UPDATE productos
                    SET 
                        nombre        = %s,
                        unidad        = %s,
                        descripcion   = %s,
                        categoria     = %s,
                        stock         = %s,
                        stock_minimo  = %s,
                        estado_stock  = %s,
                        precio_compra = %s,
                        precio_venta  = %s
                    WHERE id = %s
                """
                params = (
                    data.form['nombre'],
                    data.form['unidad'],
                    data.form['descripcion'],
                    data.form['categoria'],
                    data.form['stock'],
                    data.form['stock_minimo'],
                    data.form['estado_stock'],
                    data.form['precio_compra'],
                    data.form['precio_venta'],
                    data.form['id']  # hidden en el form
                )
                cursor.execute(query, params)
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en procesar_actualizacion_producto: {e}")
        return None


def eliminarProducto(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                query = "DELETE FROM productos WHERE id = %s"
                cursor.execute(query, (id,))
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en eliminarProducto: {e}")
        return []


def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                cursor.execute("""
                    SELECT 
                        id,
                        nombre        AS name_surname,
                        email         AS email_user,
                        fecha_registro AS created_user,
                        estado
                    FROM clientes
                    ORDER BY id DESC
                """)
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en lista_usuariosBD: {e}")
        return []


def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                cursor.execute("DELETE FROM tbl_empleados WHERE id_empleado = %s", (id_empleado,))
                conexion_MySQLdb.commit()
                deleted = cursor.rowcount

                if deleted:
                    basepath = path.dirname(__file__)
                    urlFile = path.join(basepath, '../static/fotos_empleados', foto_empleado)
                    if path.exists(urlFile):
                        remove(urlFile)

                return deleted
    except Exception as e:
        print(f"Error en eliminarEmpleado: {e}")
        return []


def eliminarUsuario(id_usuario):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "DELETE FROM clientes WHERE id = %s"
                cursor.execute(sql, (id_usuario,))
                conexion_MySQLdb.commit()
                return cursor.rowcount   # 1 si borró, 0 si no
    except Exception as e:
        print(f"Error en eliminarUsuario: {e}")
        return 0

    
# =========================
#   SERVICIOS
# =========================

def sql_lista_serviciosBD():
    """Devuelve la lista de servicios registrados."""
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        id_servicio,
                        nombre       AS nombre_servicio,
                        descripcion  AS descripcion_servicio,
                        precio       AS precio_referencial,
                        tiempo_min   AS duracion_aprox_min,
                        estado       AS estado_servicio,
                        fecha_registro
                    FROM servicios
                    ORDER BY id_servicio DESC
                """
                cursor.execute(querySQL)
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en sql_lista_serviciosBD: {e}")
        return []

def procesar_form_servicio(dataForm):
    """Inserta un nuevo servicio en la tabla servicios."""
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """
                    INSERT INTO servicios (
                        nombre,
                        descripcion,
                        precio,
                        tiempo_min,
                        estado
                    ) VALUES (%s, %s, %s, %s, %s)
                """

                valores = (
                    dataForm['nombre_servicio'],
                    dataForm['descripcion_servicio'],
                    dataForm['precio_referencial'],
                    dataForm['duracion_aprox_min'],
                    dataForm.get('estado_servicio', 'ACTIVO')
                )

                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en procesar_form_servicio: {e}")
        return None



def eliminarServicio(id_servicio):
    """Elimina un servicio por su ID."""
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                query = "DELETE FROM servicios WHERE id_servicio = %s"
                cursor.execute(query, (id_servicio,))
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en eliminarServicio: {e}")
        return None

def sql_historial_serviciosBD():
    """
    Devuelve el historial de órdenes de servicio.
    Ajusta nombres de columnas según tu diseño.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        os.id_orden_servicio       AS id_orden,
                        os.fecha_registro          AS fecha,
                        c.nombre                   AS cliente,
                        os.total_pagar             AS total,
                        os.estado_orden            AS estado
                    FROM ordenes_servicio os
                    JOIN clientes c 
                        ON c.id = os.id_cliente
                    ORDER BY os.fecha_registro DESC
                """
                cursor.execute(querySQL)
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en sql_historial_serviciosBD: {e}")
        return []


